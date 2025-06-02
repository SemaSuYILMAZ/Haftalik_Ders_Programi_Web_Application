import pyodbc
import openpyxl
import pandas as pd
import os
from django.conf import settings
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def get_connection(database=None):
    connection_string = (
        'DRIVER={ODBC Driver 17 for SQL Server};'
        'SERVER=LAPTOP-0D5M84FN\SQLEXPRESS;'
        + (f'DATABASE={database};' if database else '') +
        'Trusted_Connection=yes;'
    )
    conn = pyodbc.connect(connection_string)
    conn.autocommit = True
    return conn

# Excel oluşturma işlemleri
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ders Programı"

days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
time_slots = ["09:00-10:00", "10:00-11:00", "11:00-12:00",
              "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00",
              "16:00-17:00", "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00"]
bm_class_headers = ["1. Sınıf", "2. Sınıf", "3. Sınıf", "4. Sınıf"]

# Hücre genişliklerini ayarlama
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 15
for col in range(3, 3 + len(bm_class_headers)):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"))

# Başlıklar
ws.merge_cells("A1:B1")
ws["A1"] = "Bölüm"
ws["A1"].font = Font(bold=True)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:B2")
ws["A2"] = "Gün/Saatler"
ws["A2"].font = Font(bold=True)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("C1:F1")
ws["C1"] = "Bilgisayar Mühendisliği"
ws["C1"].font = Font(bold=True)
ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
ws["C1"].border = thin_border
ws["F1"].border = thin_border

# Sınıf başlıklarını ekleme
for idx, header in enumerate(bm_class_headers, start=3):
    cell = ws.cell(row=2, column=idx, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Gün ve saatleri ekleme
row_num = 3
for day in days:
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + len(time_slots) - 1, end_column=1)
    cell = ws.cell(row=row_num, column=1, value=day)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True)
    cell.border = thin_border

    for time_slot in time_slots:
        cell = ws.cell(row=row_num, column=2, value=time_slot)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = thin_border
        row_num += 1

colors = ["FFDDC1", "D3E3FC", "FAF4B7", "D4E2D4"]
for row in ws.iter_rows(min_row=3, max_row=row_num - 1, min_col=3, max_col=6):
    for idx, cell in enumerate(row):
        cell.fill = PatternFill(start_color=colors[idx], end_color=colors[idx], fill_type="solid")
        cell.border = thin_border

row_num += 2
sw_class_headers = ["1. Sınıf", "2. Sınıf", "3. Sınıf"]
ws.merge_cells(f"A{row_num}:B{row_num}")
ws[f"A{row_num}"] = "Bölüm"
ws[f"A{row_num}"].font = Font(bold=True)
ws[f"A{row_num}"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells(f"A{row_num + 1}:B{row_num + 1}")
ws[f"A{row_num + 1}"] = "Gün/Saatler"
ws[f"A{row_num + 1}"].font = Font(bold=True)
ws[f"A{row_num + 1}"].alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells(f"C{row_num}:E{row_num}")
ws[f"C{row_num}"] = "Yazılım Mühendisliği"
ws[f"C{row_num}"].font = Font(bold=True)
ws[f"C{row_num}"].alignment = Alignment(horizontal="center", vertical="center")
ws[f"C{row_num}"].border = thin_border

top_border = Border(top=Side(style="thin"), right=Side(style="thin"))
for col in range(1, 6):
    ws.cell(row=60, column=col).border = top_border

# Sınıf başlıklarını ekleme
for idx, header in enumerate(sw_class_headers, start=3):
    cell = ws.cell(row=row_num + 1, column=idx, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Gün ve saatleri ekleme
row_num += 2
for day in days:
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + len(time_slots) - 1, end_column=1)
    cell = ws.cell(row=row_num, column=1, value=day)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True)
    cell.border = thin_border

    for time_slot in time_slots:
        cell = ws.cell(row=row_num, column=2, value=time_slot)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        cell.border = thin_border
        row_num += 1

# Yazılım Mühendisliği için hücreleri renklendirme
for row in ws.iter_rows(min_row=row_num - len(days) * len(time_slots), max_row=row_num - 1, min_col=3, max_col=5):
    for idx, cell in enumerate(row):
        cell.fill = PatternFill(start_color=colors[idx], end_color=colors[idx], fill_type="solid")
        cell.border = thin_border

# Excel dosyasını kaydet
media_file_path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
wb.save(media_file_path)
time_slots = ["09:00-10:00", "10:00-11:00", "11:00-12:00",
              "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00",
              "16:00-17:00", "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00"]


# Online dersler
def get_online_courses():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()
    online_courses = []

    # Yanlızca online dersleri çekme işlevi
    cursor.execute("""
        SELECT ders_adi, haftalik_saat, ogrv_id, sinif, bolum_id, zorunlu_saat
        FROM CourseSchedule_dersler
        WHERE online = 'Evet'  -- SADECE ONLINE DERSLERİ AL
    """)

    online_courses = cursor.fetchall()
    conn.close()
    return online_courses


# Ortak dersleri belirleyip liste olarak döndür
def get_common_courses():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()
    common_courses = set()

    cursor.execute("""
        SELECT DISTINCT ders_adi, haftalik_saat, ogrv_id, sinif
        FROM CourseSchedule_dersler
        WHERE (online = 'Hayır' OR online IS NULL)  -- Online dersleri hariç tut
        AND ders_adi IN (
            SELECT ders_adi FROM CourseSchedule_dersler WHERE bolum_id IN (1, 2)
            GROUP BY ders_adi HAVING COUNT(DISTINCT bolum_id) = 2
        )
    """)

    for course in cursor.fetchall():
        course_name, hours_per_week, instructor_id, class_year = course
        common_courses.add((course_name, hours_per_week, instructor_id, class_year))

    conn.close()
    return list(common_courses)

#Belirtilen başlangıç ve bitiş saatleri arasındaki tüm saat aralıklarını oluşturur.
def expand_time_range(start_time, end_time):
    start_time = start_time.strip()
    end_time = end_time.strip()

    try:
        start = datetime.strptime(start_time, "%H:%M")
        end = datetime.strptime(end_time, "%H:%M")
    except ValueError as e:
        print(f"Hata: Saat formatı hatalı! start_time='{start_time}', end_time='{end_time}'")
        return []  # Hatalı veriyi boş listeyle dön

    slots = []
    while start < end:
        next_hour = start + timedelta(hours=1)
        slots.append(f"{start.strftime('%H:%M')}-{next_hour.strftime('%H:%M')}")
        start = next_hour

    return slots


def get_instructor_availability():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()
    instructor_availability = {}

    # Öğretim üyelerinin uygun saatlerini çek
    cursor.execute("""
        SELECT ogrv_id, ogretim_gorevlisi, pazartesi, sali, carsamba, persembe, cuma, kullanici_adi, sifre
        FROM CourseSchedule_ogretimgorevlileri
    """)

    rows = cursor.fetchall()
    if not rows:
        print(" Uyarı: Öğretim üyelerinin uygun saatleri veritabanından çekilemedi!")
        return None

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

    for row in rows:
        instructor_id = row[0]
        instructor_name = row[1]  # Öğretim üyesi adı
        instructor_availability[instructor_id] = {}

        for i, day in enumerate(days, start=2):
            available_hours = row[i]
            if available_hours:
                times = available_hours.split(", ")
                expanded_slots = []

                # Saat aralıklarını genişletme işlevi
                for j in range(len(times) - 1):
                    expanded_slots.extend(expand_time_range(times[j], times[j + 1]))

                instructor_availability[instructor_id][day] = expanded_slots
            else:
                instructor_availability[instructor_id][day] = []

    conn.close()
    return instructor_availability


def sort_instructors_by_availability(instructor_availability):
    sorted_availability = {}
    for instructor, days in instructor_availability.items():
        sorted_days = sorted(days.items(), key=lambda x: len(x[1]), reverse=True)
        sorted_availability[instructor] = {day: times for day, times in sorted_days}
    return sorted_availability


# Uygun saatleri biçimlendiren fonksiyon
def convert_times_to_slots(instructor_availability, time_slots):
    print("time_slots içeriği:", time_slots)

    converted_availability = {}
    for instructor, availability in instructor_availability.items():
        converted_availability[instructor] = {}

        for day, hours in availability.items():
            converted_availability[instructor][day] = []

            for i in range(len(hours) - 1):
                start_time = hours[i]
                end_time = hours[i + 1]

                # Saat aralığı uygunsa time_slots'a ekle
                for slot in time_slots:
                    slot_start, slot_end = slot.split('-')
                    if slot_start == start_time and slot_end == end_time:
                        converted_availability[instructor][day].append(slot)

    return converted_availability


#Online ve ortak dersler dışında kalan bölüme özel dersleri veritabanından çeker.
def get_department_courses():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()
    department_courses = []

    cursor.execute("""
        SELECT DISTINCT ders_adi, haftalik_saat, ogrv_id, sinif, bolum_id
        FROM CourseSchedule_dersler
        WHERE (online = 'Hayır' OR online IS NULL)  -- Online dersleri hariç tut
        AND ders_adi NOT IN (
            SELECT DISTINCT ders_adi FROM CourseSchedule_dersler WHERE bolum_id IN (1, 2)
            GROUP BY ders_adi HAVING COUNT(DISTINCT bolum_id) = 2
        )  -- Ortak dersleri hariç tut
    """)

    for course in cursor.fetchall():
        course_name, hours_per_week, instructor_id, class_year, department_id = course
        department_courses.append((course_name, hours_per_week, instructor_id, class_year, department_id))

    conn.close()
    return department_courses


# Veritabanından gelen 'zorunlu_saat' değerlerini uygun 'time_slots' formatına çevirir.
def convert_mandatory_time(mandatory_time):
    slot_mapping = {
        "09:00": "09:00-10:00", "10:00": "10:00-11:00", "11:00": "11:00-12:00",
        "12:00": "12:00-13:00", "13:00": "13:00-14:00", "14:00": "14:00-15:00",
        "15:00": "15:00-16:00", "16:00": "16:00-17:00",
        "17:00": "17:00-18:00", "18:00": "18:00-19:00",
        "19:00": "19:00-20:00", "20:00": "20:00-21:00"
    }

    converted_slots = []
    times = sorted(mandatory_time.split(", "))  # Saatleri sıralı hale getirir

    for i in range(len(times) - 1):
        start_time = times[i]
        end_time = times[i + 1]

        # Eğer saatler ardışık değilse aradaki tüm saatleri ekler
        while start_time in slot_mapping and start_time != end_time:
            converted_slots.append(slot_mapping[start_time])
            hours, minutes = map(int, start_time.split(":"))
            hours += 1
            start_time = f"{hours:02d}:00"  # Yeni saati oluştur

    # Son saati de ekle
    if times[-1] in slot_mapping and slot_mapping[times[-1]] not in converted_slots:
        converted_slots.append(slot_mapping[times[-1]])

    return converted_slots


def get_instructor_name(instructor_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT ogretim_gorevlisi FROM CourseSchedule_ogretimgorevlileri WHERE ogrv_id = ?
    """, (instructor_id,))
    row = cursor.fetchone()
    conn.close()

    return row[0] if row else "Bilinmeyen Öğretim Üyesi"


def assign_courses_to_schedule(online_courses, time_slots):
    media_file_path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
    wb = openpyxl.load_workbook(media_file_path)
    ws = wb.active

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
    # Boş program tablosu oluştur
    schedule = {
        day: {slot: {"Yazılım Mühendisliği": {1: None, 2: None, 3: None},
                     "Bilgisayar Mühendisliği": {1: None, 2: None, 3: None, 4: None}}
              for slot in time_slots}
        for day in days
    }

    # Dersin kaç saat atandığını takip eden sözlük işlevi
    assigned_hours_per_course = {course[0]: 0 for course in online_courses}

    # Öncelikle en fazla saat gerektiren dersleri sıralayan işlev (Büyükten küçüğe)
    online_courses.sort(key=lambda x: x[1], reverse=True)

    for course in online_courses:
        course_name, hours_per_week, instructor_id, class_year, department_id, mandatory_time = course
        instructor_name = get_instructor_name(instructor_id)

        # Eğer ders zaten atanmış ve haftalık saat dolmuşsa atamayı geç
        if assigned_hours_per_course[course_name] >= hours_per_week:
            continue

        assigned_hours = assigned_hours_per_course[course_name]
        assigned_slots = []

        # Dersin ait olduğu bölüm belirlenir
        department = "Bilgisayar Mühendisliği" if department_id == 2 else "Yazılım Mühendisliği"

        # Eğer ders her iki bölümde de aynı sınıfta okutuluyorsa bunu işaretle
        is_shared = any(c[0] == course_name and c[3] == class_year and c[4] != department_id for c in online_courses)

        # Zorunlu saatleri uygun formata çevir
        valid_mandatory_slots = convert_mandatory_time(mandatory_time)

        # Öncelikle zorunlu saatleri yerleştir
        for selected_day in days:
            for slot in valid_mandatory_slots:
                if assigned_hours >= hours_per_week:
                    break

                if schedule[selected_day][slot][department].get(class_year) is None:
                    schedule[selected_day][slot][department][class_year] = f"{course_name}\n{instructor_name}"
                    assigned_hours += 1
                    assigned_hours_per_course[course_name] = assigned_hours
                    assigned_slots.append(f"{selected_day}, {slot}, {class_year}. sınıf - {department}")

                    if is_shared:
                        other_department = "Bilgisayar Mühendisliği" if department == "Yazılım Mühendisliği" else "Yazılım Mühendisliği"
                        if schedule[selected_day][slot][other_department].get(class_year) is None:
                            schedule[selected_day][slot][other_department][
                                class_year] = f"{course_name}\n{instructor_name}"

        # Eğer hala boş saatler varsa kalanları yerleştir
        for selected_day in days:
            for slot in time_slots:
                if assigned_hours >= hours_per_week:
                    break

                if schedule[selected_day][slot][department].get(class_year) is None:
                    schedule[selected_day][slot][department][class_year] = course_name
                    assigned_hours += 1
                    assigned_hours_per_course[course_name] = assigned_hours
                    assigned_slots.append(f"{selected_day}, {slot}, {class_year}. sınıf - {department}")

                    if is_shared:
                        other_department = "Bilgisayar Mühendisliği" if department == "Yazılım Mühendisliği" else "Yazılım Mühendisliği"
                        if schedule[selected_day][slot][other_department].get(class_year) is None:
                            schedule[selected_day][slot][other_department][
                                class_year] = f"{course_name}\n{instructor_name}(Online)"

    # 4. Sınıfa 3. Sınıfın Derslerini Kopyala
    for day, slots in schedule.items():
        for slot, departments in slots.items():
            if departments["Bilgisayar Mühendisliği"][4] is None and departments["Bilgisayar Mühendisliği"][
                3] is not None:
                departments["Bilgisayar Mühendisliği"][4] = departments["Bilgisayar Mühendisliği"][3]

    # Debug için terminalde schedule'ı yazdır
    '''print("\n Debug İçin: Atama Sonrası Schedule Kontrolü\n")
    for day, slots in schedule.items():
        print(f" {day}:")
        for time_slot, classes in slots.items():
            for department, class_data in classes.items():
                for class_year, course in class_data.items():
                    if course:
                        print(f"   {time_slot} | {class_year}. sınıf - {department} -> {course}")'''

    wb.save(media_file_path)

    # Dersleri excel'e yaz
    row_offsets = {
        "Pazartesi": 3,
        "Salı": 15,
        "Çarşamba": 27,
        "Perşembe": 39,
        "Cuma": 51
    }

    sw_row_offsets = {
        "Pazartesi": 67,
        "Salı": 79,
        "Çarşamba": 91,
        "Perşembe": 103,
        "Cuma": 115
    }

    class_columns_bm = {1: 3, 2: 4, 3: 5, 4: 6}  # Bilgisayar Mühendisliği için sınıf sütunları
    class_columns_sw = {1: 3, 2: 4, 3: 5}  # Yazılım Mühendisliği için sınıf sütunları

    for day, slots in schedule.items():
        for slot, classes in slots.items():
            try:
                slot_index = time_slots.index(slot)
            except ValueError:
                continue  # Eğer zaman dilimi bulunamazsa atla

            # Bilgisayar Mühendisliği'ni yaz
            row_num_bm = row_offsets[day] + slot_index
            for class_year, course_name in classes["Bilgisayar Mühendisliği"].items():
                if course_name and class_year in class_columns_bm:
                    col = class_columns_bm[class_year]
                    ws.cell(row=row_num_bm, column=col, value=f"{course_name}\n(Online)")
                    ws.cell(row=row_num_bm, column=col).alignment = Alignment(wrapText=True)

            # Yazılım Mühendisliği'ni yaz
            row_num_sw = sw_row_offsets[day] + slot_index
            for class_year, course_name in classes["Yazılım Mühendisliği"].items():
                if course_name and class_year in class_columns_sw:
                    col = class_columns_sw[class_year]
                    ws.cell(row=row_num_sw, column=col, value=f"{course_name}\n(Online)")
                    ws.cell(row=row_num_sw, column=col).alignment = Alignment(wrapText=True)

    wb.save(media_file_path)


def assign_common_courses(common_courses, instructor_availability, time_slots):
    media_file_path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
    wb = openpyxl.load_workbook(media_file_path)
    ws = wb.active
    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

    # Mevcut ders programını çek
    schedule = {
        day: {slot: {"Yazılım Mühendisliği": {1: None, 2: None, 3: None},
                     "Bilgisayar Mühendisliği": {1: None, 2: None, 3: None, 4: None}}
              for slot in time_slots}
        for day in days
    }

    # Mevcut excelden dersleri al ve schedule'a kaydet
    for day, row_offset in zip(days, [3, 15, 27, 39, 51]):
        for slot_index, slot in enumerate(time_slots):
            row_bm = row_offset + slot_index
            row_sw = row_offset + 64 + slot_index

            for class_year in range(1, 5):
                bm_col = {1: 3, 2: 4, 3: 5, 4: 6}.get(class_year, None)
                sw_col = {1: 3, 2: 4, 3: 5}.get(class_year, None)

                if bm_col:
                    existing_bm_course = ws.cell(row=row_bm, column=bm_col).value
                    if existing_bm_course:
                        schedule[day][slot]["Bilgisayar Mühendisliği"][class_year] = existing_bm_course

                if sw_col:
                    existing_sw_course = ws.cell(row=row_sw, column=sw_col).value
                    if existing_sw_course:
                        schedule[day][slot]["Yazılım Mühendisliği"][class_year] = existing_sw_course

    # Öğretim üyesinin uygun saatlerinin belirlenmesi
    instructor_schedule = {instructor: {day: [] for day in days} for instructor in instructor_availability}

    # Sözlük ile derslerin kaç saat atandığını kontrol etme
    assigned_hours_per_course = {course[0]: 0 for course in common_courses}

    # Ortak dersleri öncelik sırasına göre ekleme
    common_courses.sort(key=lambda x: (x[3], -x[1]))

    for course in common_courses:
        course_name, hours_per_week, instructor_id, class_year = course
        instructor_name = get_instructor_name(instructor_id)

        if instructor_id not in instructor_availability:
            print(f" Öğretim Üyesi ID {instructor_id} için uygunluk bilgisi bulunamadı. {course_name} atlanıyor.")
            continue

        if assigned_hours_per_course[course_name] >= hours_per_week:
            continue

        assigned_hours = 0
        assigned_slots = []

        # Ortak dersi alan sınıfların bulunması
        related_classes = [c[3] for c in common_courses if c[0] == course_name]

        best_block = None
        for selected_day in days:
            if selected_day not in instructor_availability[instructor_id]:
                continue

            available_slots = instructor_availability[instructor_id][selected_day]
            # Boş olan saaatlerin bulunması
            for i in range(len(available_slots) - (hours_per_week - 1)):
                block_slots = available_slots[i:i + hours_per_week]

                is_valid = all([
                    slot in time_slots for slot in block_slots
                ]) and all([
                    slot not in instructor_schedule[instructor_id][selected_day] for slot in block_slots
                ]) and all([
                    all(schedule[selected_day][slot]["Bilgisayar Mühendisliği"].get(cls) is None and
                        schedule[selected_day][slot]["Yazılım Mühendisliği"].get(cls) is None
                        for cls in related_classes)
                    for slot in block_slots
                ])

                if is_valid:
                    best_block = (selected_day, block_slots)
                    break  # Uygun ilk blok bulunduğunda döngüden çıkılır

        # Eğer en iyi blok bulunduysa ders atanır
        if best_block:
            selected_day, block_slots = best_block

            for slot in block_slots:
                for cls in related_classes:
                    schedule[selected_day][slot]["Bilgisayar Mühendisliği"][cls] = f"{course_name}\n{instructor_name}"
                    schedule[selected_day][slot]["Yazılım Mühendisliği"][cls] = f"{course_name}\n{instructor_name}"

                instructor_schedule[instructor_id][selected_day].append(slot)
                assigned_hours += 1
                assigned_hours_per_course[course_name] = assigned_hours
                assigned_slots.append(f"{selected_day}, {slot}, {related_classes} sınıfları - Ortak Ders")

    for day, row_offset in zip(days, [3, 15, 27, 39, 51]):
        for slot_index, slot in enumerate(time_slots):
            row_bm = row_offset + slot_index
            row_sw = row_offset + 64 + slot_index

            for class_year in range(1, 5):
                bm_col = {1: 3, 2: 4, 3: 5, 4: 6}.get(class_year, None)
                sw_col = {1: 3, 2: 4, 3: 5}.get(class_year, None)

                if bm_col and schedule[day][slot]["Bilgisayar Mühendisliği"][class_year]:
                    course_name = schedule[day][slot]["Bilgisayar Mühendisliği"][class_year]
                    ws.cell(row=row_bm, column=bm_col, value=f"{course_name}\n")
                    ws.cell(row=row_bm, column=bm_col).alignment = Alignment(wrapText=True)

                if sw_col and schedule[day][slot]["Yazılım Mühendisliği"][class_year]:
                    course_name = schedule[day][slot]["Yazılım Mühendisliği"][class_year]
                    ws.cell(row=row_sw, column=sw_col, value=f"{course_name}\n")
                    ws.cell(row=row_sw, column=sw_col).alignment = Alignment(wrapText=True)

    wb.save(media_file_path)


#Bölüme özel dersleri uygun boş saatlere yerleştirir ve excel'e kaydeder.
def assign_department_courses(department_courses, instructor_availability, time_slots):
    media_file_path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
    wb = openpyxl.load_workbook(media_file_path)
    ws = wb.active

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

    # Mevcut ders programını çek (Online ve ortak derslerin olduğu program)
    schedule = {
        day: {slot: {"Yazılım Mühendisliği": {1: None, 2: None, 3: None},
                     "Bilgisayar Mühendisliği": {1: None, 2: None, 3: None, 4: None}}
              for slot in time_slots}
        for day in days
    }

    # Mevcut excelden dersleri al ve schedule'a kaydet
    for day, row_offset in zip(days, [3, 15, 27, 39, 51]):
        for slot_index, slot in enumerate(time_slots):
            row_bm = row_offset + slot_index
            row_sw = row_offset + 64 + slot_index

            for class_year in range(1, 5):
                bm_col = {1: 3, 2: 4, 3: 5, 4: 6}.get(class_year, None)
                sw_col = {1: 3, 2: 4, 3: 5}.get(class_year, None)

                if bm_col:
                    existing_bm_course = ws.cell(row=row_bm, column=bm_col).value
                    if existing_bm_course:
                        schedule[day][slot]["Bilgisayar Mühendisliği"][class_year] = existing_bm_course

                if sw_col:
                    existing_sw_course = ws.cell(row=row_sw, column=sw_col).value
                    if existing_sw_course:
                        schedule[day][slot]["Yazılım Mühendisliği"][class_year] = existing_sw_course

    # Öğretim üyelerinin uygun saatlerini alır
    instructor_schedule = {instructor: {day: [] for day in days} for instructor in instructor_availability}

    # Belirtilen öğretim üyesinin bu saat diliminde başka sınıfta dersi olup olmadığını kontrol eder.
    def is_instructor_available(instructor_name, selected_day, slot):
        for other_department in ["Bilgisayar Mühendisliği", "Yazılım Mühendisliği"]:
            for other_class in range(1, 5):
                assigned_course = schedule[selected_day][slot][other_department].get(other_class)
                if assigned_course and instructor_name in assigned_course:
                    return False  # Eğitmen bu saatte uygun değil
        return True  # Eğitmen bu saatte uygun

    # Bölüm derslerini öncelik sırasına göre sırala (Saat sayısına göre büyükten küçüğe)
    department_courses.sort(key=lambda x: x[1], reverse=True)

    for course in department_courses:
        course_name, hours_per_week, instructor_id, class_year, department_id = course
        instructor_name = get_instructor_name(instructor_id)

        if instructor_id not in instructor_availability:
            print(f"Öğretim Üyesi ID {instructor_id} için uygunluk bilgisi bulunamadı. {course_name} atlanıyor.")
            continue

        assigned_hours = 0
        assigned_slots = []
        department = "Bilgisayar Mühendisliği" if department_id == 2 else "Yazılım Mühendisliği"

        best_block = None
        for selected_day in days:
            available_slots = instructor_availability.get(instructor_id, {}).get(selected_day, [])

            for i in range(len(available_slots) - (hours_per_week - 1)):
                block_slots = available_slots[i:i + hours_per_week]
                is_valid = all([
                    slot in time_slots for slot in block_slots
                ]) and all([
                    is_instructor_available(instructor_name, selected_day, slot) for slot in block_slots
                ]) and all([
                    schedule[selected_day][slot][department].get(class_year) is None for slot in block_slots
                ])

                if is_valid:
                    best_block = (selected_day, block_slots)
                    break

        if best_block:
            selected_day, block_slots = best_block
            for slot in block_slots:
                schedule[selected_day][slot][department][class_year] = f"{course_name}\n{instructor_name}"
                instructor_schedule[instructor_id][selected_day].append(slot)
                assigned_hours += 1
                assigned_slots.append(f"{selected_day}, {slot}, {class_year}. sınıf - {department}")

    for day, row_offset in zip(days, [3, 15, 27, 39, 51]):
        for slot_index, slot in enumerate(time_slots):
            row_bm = row_offset + slot_index
            row_sw = row_offset + 64 + slot_index

            for class_year in range(1, 5):
                bm_col = {1: 3, 2: 4, 3: 5, 4: 6}.get(class_year, None)
                sw_col = {1: 3, 2: 4, 3: 5}.get(class_year, None)

                if bm_col and schedule[day][slot]["Bilgisayar Mühendisliği"][class_year]:
                    course_name = schedule[day][slot]["Bilgisayar Mühendisliği"][class_year]
                    ws.cell(row=row_bm, column=bm_col, value=f"{course_name}\n")
                    ws.cell(row=row_bm, column=bm_col).alignment = Alignment(wrapText=True)

                if sw_col and schedule[day][slot]["Yazılım Mühendisliği"][class_year]:
                    course_name = schedule[day][slot]["Yazılım Mühendisliği"][class_year]
                    ws.cell(row=row_sw, column=sw_col, value=f"{course_name}\n")
                    ws.cell(row=row_sw, column=sw_col).alignment = Alignment(wrapText=True)

    wb.save(media_file_path)


# Derslikler
def get_classrooms():
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT derslik_id, kapasite, statu FROM CourseSchedule_derslikler
    """)

    classrooms = []
    for row in cursor.fetchall():
        classrooms.append({
            "class_id": row[0],
            "capacity": row[1],
            "status": row[2]
        })

    conn.close()
    return classrooms


# Dersi alan öğrenci sayısı
def get_student_count_for_course(course_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT COUNT(*) FROM CourseSchedule_ogrenciders WHERE course_id = ?
    """, (course_id,))

    row = cursor.fetchone()
    conn.close()

    return row[0] if row else 0


# Dersler
def get_course_id(course_name):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT course_id FROM CourseSchedule_dersler WHERE ders_adi = ?
    """, (course_name,))

    row = cursor.fetchone()
    conn.close()

    return row[0] if row else None


# Online dersler
def get_online_status(course_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("SELECT online FROM CourseSchedule_dersler WHERE course_id = ?", (course_id,))
    row = cursor.fetchone()
    conn.close()

    return row[0]

media_file_path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
# Kayıtlı excel dosyası
def read_courses_from_excel(filename=media_file_path):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    schedule = {}

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
    row_offsets = {"Pazartesi": 3, "Salı": 15, "Çarşamba": 27, "Perşembe": 39, "Cuma": 51}
    sw_row_offsets = {"Pazartesi": 67, "Salı": 79, "Çarşamba": 91, "Perşembe": 103, "Cuma": 115}

    class_columns_bm = {1: 3, 2: 4, 3: 5, 4: 6}
    class_columns_sw = {1: 3, 2: 4, 3: 5}

    for day in days:
        schedule[day] = {}
        for slot in time_slots:
            schedule[day][slot] = {
                "Bilgisayar Mühendisliği": {1: None, 2: None, 3: None, 4: None},
                "Yazılım Mühendisliği": {1: None, 2: None, 3: None}
            }

            slot_index = time_slots.index(slot)
            row_bm = row_offsets[day] + slot_index
            row_sw = sw_row_offsets[day] + slot_index

            for class_year in range(1, 5):
                bm_col = class_columns_bm.get(class_year, None)
                if bm_col:
                    cell_value = ws.cell(row=row_bm, column=bm_col).value
                    if cell_value:
                        schedule[day][slot]["Bilgisayar Mühendisliği"][class_year] = cell_value.strip()

            for class_year in range(1, 4):
                sw_col = class_columns_sw.get(class_year, None)
                if sw_col:
                    cell_value = ws.cell(row=row_sw, column=sw_col).value
                    if cell_value:
                        schedule[day][slot]["Yazılım Mühendisliği"][class_year] = cell_value.strip()

    return schedule


# Ders statülerini belirleme
def get_course_status(course_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("""
        SELECT statu FROM CourseSchedule_dersler WHERE course_id = ?
    """, (course_id,))

    row = cursor.fetchone()
    conn.close()

    return row[0] if row and row[0] else "NORMAL"


# Haftalık toplam ders saati
def get_course_duration(course_id):
    conn = get_connection(database='DersProgramiDB')
    cursor = conn.cursor()

    cursor.execute("SELECT haftalik_saat FROM CourseSchedule_dersler WHERE course_id = ?", (course_id,))
    row = cursor.fetchone()
    conn.close()

    return int(row[0])


# Derslik ataması
def assign_classrooms_to_courses(schedule, time_slots):
    classrooms = get_classrooms()
    course_classroom_map = {}  # Her dersin ilk atanan dersliğini tutar
    occupied_classrooms = {}  # Hangi dersliklerin dolu olduğunu tutar
    course_duration_map = {}  # Hangi dersin kaç saat sürdüğünü tutar

    for day, slots in schedule.items():
        for slot_index, slot in enumerate(time_slots):
            for department, classes in schedule[day][slot].items():
                for class_year, course_info in classes.items():
                    if not course_info:
                        continue

                    lines = course_info.split("\n")
                    course_name = lines[0]
                    instructor_name = lines[1] if len(lines) > 1 else "Bilinmeyen Eğitmen"

                    # Dersin ID'si
                    course_id = get_course_id(course_name)
                    if not course_id:
                        print(f"⚠ {course_name} için ders bulunamadı, ancak derslik ataması devam ediyor.")
                        continue

                    # Ders online mı kontrolü
                    online_status = get_online_status(course_id)
                    if online_status == 'Evet':
                        schedule[day][slot][department][class_year] = f"{course_name}\n{instructor_name} (Online)"
                        continue

                    # Dersin kaç saat olduğu bilgisi
                    if course_name not in course_duration_map:
                        course_duration_map[course_name] = get_course_duration(course_id)
                    duration = course_duration_map[course_name]

                    if course_name in course_classroom_map:
                        classroom_id = course_classroom_map[course_name]
                    else:
                        course_status = get_course_status(course_id)
                        student_count = get_student_count_for_course(course_id)  #Dersin öğrenci sayısı

                        # Ders için uygun kapasitedeki derslikleri filtreleme
                        suitable_classrooms = sorted(
                            [c for c in classrooms if c["capacity"] >= student_count and c["status"] == course_status],
                            key=lambda x: x["capacity"])

                        if not suitable_classrooms:
                            print(f"{course_name} için uygun derslik bulunamadı!")
                            continue

                        # İlk uygun ve müsait dersliği seç
                        classroom_id = None
                        for classroom in suitable_classrooms:
                            # Eğer bu derslik başka bir dersin saatleri içinde doluysa geç
                            is_available = True
                            for i in range(duration):  # Dersin süresi boyunca kontrol et
                                future_slot = time_slots[slot_index + i] if slot_index + i < len(time_slots) else None
                                if future_slot and (day, future_slot) in occupied_classrooms and classroom["class_id"] in \
                                        occupied_classrooms[(day, future_slot)]:
                                    is_available = False
                                    break

                            if is_available:
                                classroom_id = classroom["class_id"]
                                course_classroom_map[course_name] = classroom_id
                                break
                        else:
                            continue

                    # Dersliğin bu gün ve saatte dolu olduğunu kaydet
                    for i in range(duration):
                        future_slot = time_slots[slot_index + i] if slot_index + i < len(time_slots) else None
                        if future_slot:
                            if (day, future_slot) not in occupied_classrooms:
                                occupied_classrooms[(day, future_slot)] = set()
                            occupied_classrooms[(day, future_slot)].add(classroom_id)

                    # Programda dersi güncelle ve dersliği ekle
                    schedule[day][slot][department][class_year] = f"{course_name}\n{instructor_name} ({classroom_id})"

    return schedule


def generate_schedule_excel(excel_path):
    print("\n📌 Dersler veritabanından çekiliyor.")
    online_courses = get_online_courses()   #Online dersler veritabanından çekiliyor
    common_courses = get_common_courses()  #Ortak dersler veritabanından çekiliyor
    department_courses = get_department_courses() #Bölüme özel dersler veritabanından çekiliyor

    print("📌 Öğretim üyelerinin uygunluk durumu alınıyor.")
    instructor_availability = get_instructor_availability()
    if instructor_availability is None:
        print("Hata: Öğretim üyesi uygunluk verisi çekilemedi. Veritabanını kontrol et!")
        exit(1)

    time_slots = [
        "09:00-10:00", "10:00-11:00", "11:00-12:00",
        "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00",
        "16:00-17:00", "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00"
    ]

    print("📌 Ders programı oluşturuluyor ve Excel'e kaydediliyor.")

    print("\n📌 Dersler Atanıyor.")
    # Online Dersleri Atama İşlevi
    if online_courses:
        assign_courses_to_schedule(online_courses, time_slots)
    else:
        print("Atanacak online ders bulunamadı!")

    # Ortak Dersleri Atama İşlevi
    if common_courses:
        assign_common_courses(common_courses, instructor_availability, time_slots)
    else:
        print("Atanacak ortak ders bulunamadı!")

    # Bölüme Özel Dersleri Atama İşlevi
    if department_courses:
        assign_department_courses(department_courses, instructor_availability, time_slots)
    else:
        print("Atanacak bölüme özel ders bulunamadı!")

    # Kaydedilen Excel'den Dersleri Okuma İşlevi
    schedule = read_courses_from_excel()

    # Derslik Ataması Yapma İşlevi
    print("📌 Derslikler Atanıyor.")
    schedule = assign_classrooms_to_courses(schedule, time_slots)

    # Güncellenmiş Programı Excel'e Kaydet
    print("\n✅ Ders Programı Excel'e Kaydediliyor.")
    media_file_path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
    wb = openpyxl.load_workbook(media_file_path)
    ws = wb.active

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
    row_offsets = {"Pazartesi": 3, "Salı": 15, "Çarşamba": 27, "Perşembe": 39, "Cuma": 51}
    sw_row_offsets = {"Pazartesi": 67, "Salı": 79, "Çarşamba": 91, "Perşembe": 103, "Cuma": 115}

    class_columns_bm = {1: 3, 2: 4, 3: 5, 4: 6}
    class_columns_sw = {1: 3, 2: 4, 3: 5}

    for day, slots in schedule.items():
        for slot, classes in slots.items():
            try:
                slot_index = time_slots.index(slot)
            except ValueError:
                continue

            row_num_bm = row_offsets[day] + slot_index
            row_num_sw = sw_row_offsets[day] + slot_index

            # Bilgisayar Mühendisliği'ni yaz
            for class_year, course_name in classes["Bilgisayar Mühendisliği"].items():
                if course_name and class_year in class_columns_bm:
                    ws.cell(row=row_num_bm, column=class_columns_bm[class_year], value=f"{course_name}\n")
                    ws.cell(row=row_num_bm, column=class_columns_bm[class_year]).alignment = Alignment(wrapText=True)

            # Yazılım Mühendisliği'ni yaz
            for class_year, course_name in classes["Yazılım Mühendisliği"].items():
                if course_name and class_year in class_columns_sw:
                    ws.cell(row=row_num_sw, column=class_columns_sw[class_year], value=f"{course_name}\n")
                    ws.cell(row=row_num_sw, column=class_columns_sw[class_year]).alignment = Alignment(wrapText=True)

    wb.save(excel_path)


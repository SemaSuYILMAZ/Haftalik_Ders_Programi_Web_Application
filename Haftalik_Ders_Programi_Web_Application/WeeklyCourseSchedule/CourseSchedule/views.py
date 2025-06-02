import os
import json
import openpyxl
from django.db import connection
from django.conf import settings
from .forms import ManuelAtamaForm
from openpyxl import load_workbook
from django.contrib import messages
from django.db.models import F, Value
from datetime import datetime, timedelta
from .main import generate_schedule_excel
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter
from collections import defaultdict, OrderedDict
from django.db.models.functions import Concat, Trim
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, FileResponse, HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from .models import Ogrenciler, Dersler, OgretimGorevlileri, Derslikler, OgrenciDers


def index(request):
    context = {}

    if request.method == 'POST':
        # Öğrenci girişi
        if 'student_username' in request.POST:
            no = request.POST['student_username']
            sifre = request.POST['student_password']
            try:
                ogrenci = Ogrenciler.objects.get(numara=no, sifre=sifre)
                request.session['student_no'] = ogrenci.numara  # ➤ Numara oturuma kaydedildi
                return redirect('ogrenci_sayfa')  # başarılı girişte yönlendir
            except Ogrenciler.DoesNotExist:
                context['ogrenci_hata'] = "Hatalı öğrenci girişi"

        # Öğretim görevlisi girişi
        elif 'instructor_username' in request.POST:
            kullanici = request.POST['instructor_username']
            sifre = request.POST['instructor_password']
            try:
                ogretim = OgretimGorevlileri.objects.get(kullanici_adi=kullanici, sifre=sifre)
                request.session['ogretmen_adi'] = ogretim.ogretim_gorevlisi  # isim oturuma ekleniyor
                return redirect('ogretim_sayfa')  # başarılı girişte yönlendir
            except OgretimGorevlileri.DoesNotExist:
                context['ogretim_hata'] = "Hatalı öğretim görevlisi girişi"

    return render(request, 'CourseSchedule/index.html', context)


def ogrenci_sayfa(request):
    dersler = Dersler.objects.annotate(
        kod=Trim(F('ders_kodu')),
        ad=Trim(F('ders_adi'))
    ).values('kod', 'ad').distinct()

    selected_ids = request.session.get('secilen_dersler', [])
    print("Oturumdaki seçili dersler:", selected_ids)

    ogrenci_programi = []

    if selected_ids:
        selected_adlar = [d.split(" - ", 1)[-1].strip().lower() for d in selected_ids]

        dosya_yolu = os.path.join(settings.MEDIA_ROOT, "Ders_Programi.xlsx")
        wb = openpyxl.load_workbook(dosya_yolu)
        sheet = wb.active

        merged_cells_map = {}
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_cells_map[(row, col)] = top_left_value

        gun_sirasi = {"Pazartesi": 1, "Salı": 2, "Çarşamba": 3, "Perşembe": 4, "Cuma": 5}
        gorulen_satirlar = set()

        for row in sheet.iter_rows(min_row=3):
            gun = merged_cells_map.get((row[0].row, row[0].column), row[0].value)
            saat = merged_cells_map.get((row[1].row, row[1].column), row[1].value)

            if not gun or not saat:
                continue

            for cell in row[2:]:
                coord = (cell.row, cell.column)
                value = merged_cells_map.get(coord, cell.value)

                if isinstance(value, str):
                    ders_adi_clean = value.strip().split("\n")[0].strip().lower()
                    if ders_adi_clean in selected_adlar:
                        kayit = (gun, saat, value.strip())
                        if kayit not in gorulen_satirlar:
                            ogrenci_programi.append(kayit)
                            gorulen_satirlar.add(kayit)

        ogrenci_programi.sort(key=lambda x: (gun_sirasi.get(x[0], 99), str(x[1])))

    return render(request, "CourseSchedule/ogrenci.html", {
        'dersler': dersler,
        'selected_ids': selected_ids,
        'student_no': request.session.get('student_no'),
        'ogrenci_programi': ogrenci_programi,
    })


def cikis_yap(request):
    request.session.flush()  # Oturumu temizle
    return redirect('index')  # Anasayfaya yönlendir


@csrf_exempt
def ders_secimi_kaydet(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            secilen_dersler = data.get('dersler', [])
            request.session['secilen_dersler'] = secilen_dersler
            print("Seçilen dersler oturuma kaydedildi:", secilen_dersler)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            print("Hata:", e)
            return JsonResponse({'status': 'error'}, status=400)


@csrf_exempt
def ders_sil(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        silinecek_ders = data.get('ders')

        selected = request.session.get('secilen_dersler', [])
        if silinecek_ders in selected:
            selected.remove(silinecek_ders)
            request.session['secilen_dersler'] = selected
            print("Ders kaldırıldı:", silinecek_ders)
        return JsonResponse({'status': 'success'})


def ogretim_anasayfa(request):
    ogretmen_adi = request.session.get("ogretmen_adi", "Misafir")
    ogretim_uyeleri = OgretimGorevlileri.objects.only("ogretim_gorevlisi")

    aktif_menu = ""
    filtrelenmis_tablo = []
    yazilim_baslik = []
    yazilim_siniflar = []
    yazilim_satirlar = []
    secili_uye = ""
    secili_bolum = ""
    secili_sinif = ""
    sinif_programi = []

    if request.method == "POST":
        aktif_menu = request.POST.get("aktif_menu", "")
        secili_uye = request.POST.get("secili_uye", "")
        secili_bolum = request.POST.get("secili_bolum", "")
        secili_sinif = request.POST.get("secili_sinif", "")

        dosya_yolu = os.path.join(settings.MEDIA_ROOT, "Ders_Programi.xlsx")
        wb = openpyxl.load_workbook(dosya_yolu)
        sheet = wb.active

        if aktif_menu == "ogretim_uyeleri" and secili_uye:
            merged_cells_map = {}
            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                top_left_value = sheet.cell(row=min_row, column=min_col).value
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_cells_map[(row, col)] = top_left_value

            for row_index, row in enumerate(sheet.iter_rows()):
                yeni_satir = []
                for col_index, cell in enumerate(row):
                    coord = (cell.row, cell.column)
                    value = merged_cells_map.get(coord, cell.value)

                    if row_index < 2 or col_index < 2:
                        yeni_satir.append(value)
                    elif isinstance(value, str) and (
                            secili_uye.lower() in value.lower()
                            or "yazılım mühendisliği" in value.lower()
                            or "bilgisayar mühendisliği" in value.lower()
                    ):
                        yeni_satir.append(value)
                    else:
                        yeni_satir.append("")
                filtrelenmis_tablo.append(yeni_satir)

            for i, row in enumerate(filtrelenmis_tablo):
                if any(cell and "yazılım mühendisliği" in str(cell).lower() for cell in row):
                    yazilim_baslik = row
                    if i + 1 < len(filtrelenmis_tablo):
                        yazilim_siniflar = filtrelenmis_tablo[i + 1]
                        # Eğer sınıf başlıkları eksikse manuel olarak ekleyelim
                        yazilim_siniflar[0] = "Gün/Saatler"
                        yazilim_siniflar[2] = "1. Sınıf"
                        yazilim_siniflar[3] = "2. Sınıf"
                        yazilim_siniflar[4] = "3. Sınıf"
                    yazilim_satirlar = filtrelenmis_tablo[i + 2:]
                    break


        elif aktif_menu == "sinif_programi" and secili_bolum and secili_sinif:
            bolum_adi = "Yazılım Mühendisliği" if secili_bolum == "yazilim" else "Bilgisayar Mühendisliği"
            bolum_satiri_index = None
            bir_sonraki_bolum_satiri = None
            sinif_sutun_index = None

            # Bölüme ait satırları bul
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if bolum_satiri_index is None and bolum_adi in row:
                    bolum_satiri_index = i
                elif bolum_satiri_index is not None and any(
                        cell and "mühendisliği" in str(cell).lower() for cell in row):
                    bir_sonraki_bolum_satiri = i
                    break

            if bolum_satiri_index is not None:
                sinif_satiri = \
                list(sheet.iter_rows(min_row=bolum_satiri_index + 2, max_row=bolum_satiri_index + 2, values_only=True))[
                    0]
                for idx, cell in enumerate(sinif_satiri):
                    if secili_sinif in str(cell):
                        sinif_sutun_index = idx
                        break

                if sinif_sutun_index is not None:
                    onceki_gun = ""
                    for i, row in enumerate(sheet.iter_rows(min_row=bolum_satiri_index + 3, values_only=True),
                                            start=bolum_satiri_index + 3):
                        if bir_sonraki_bolum_satiri and i >= bir_sonraki_bolum_satiri:
                            break

                        gun = str(row[0]).strip() if row[0] else onceki_gun
                        saat = str(row[1]).strip() if row[1] else ""
                        ders = str(row[sinif_sutun_index]).strip() if row[sinif_sutun_index] else ""

                        if gun and saat:
                            sinif_programi.append([gun, saat, ders])
                            onceki_gun = gun  # Son geçerli günü hatırla

    return render(request, "CourseSchedule/ogretim_sayfa.html", {
        "ogretmen_adi": ogretmen_adi,
        "ogretim_uyeleri": ogretim_uyeleri,
        "filtrelenmis_tablo": filtrelenmis_tablo,
        "secili_uye": secili_uye,
        "yazilim_baslik": yazilim_baslik,
        "yazilim_siniflar": yazilim_siniflar,
        "yazilim_satirlar": yazilim_satirlar,
        "aktif_menu": aktif_menu,
        "secili_bolum": secili_bolum,
        "secili_sinif": secili_sinif,
        "sinif_programi": sinif_programi,
        "secili_bolum_ad": "Yazılım Mühendisliği" if secili_bolum == "yazilim" else "Bilgisayar Mühendisliği",

    })


def normalize_table(table):
    last_day = None
    new_table = []
    for row in table:
        row = list(row)
        if row[0] is not None:
            last_day = row[0]
        else:
            row[0] = last_day
        new_table.append(row)
    return new_table


def program_olustur_view(request):
    # Excel oluşturma işlemleri
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ders Programı"

    days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
    time_slots = ["09:00-10:00", "10:00-11:00", "11:00-12:00",
                  "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00",
                  "16:00-17:00", "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00"]
    bm_class_headers = ["1. Sınıf", "2. Sınıf", "3. Sınıf", "4. Sınıf"]

    # Hücre genişliklerini ayarlama
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    for col in range(3, 3 + len(bm_class_headers)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"))

    # Başlıklar
    ws.merge_cells("A1:B1")
    ws["A1"] = "Bölüm"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:B2")
    ws["A2"] = "Gün/Saatler"
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("C1:F1")
    ws["C1"] = "Bilgisayar Mühendisliği"
    ws["C1"].font = Font(bold=True)
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C1"].border = thin_border
    ws["F1"].border = thin_border

    # Sınıf başlıklarını ekleme
    for idx, header in enumerate(bm_class_headers, start=3):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Gün ve saatleri ekleme
    row_num = 3
    for day in days:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + len(time_slots) - 1, end_column=1)
        cell = ws.cell(row=row_num, column=1, value=day)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = thin_border

        for time_slot in time_slots:
            cell = ws.cell(row=row_num, column=2, value=time_slot)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.border = thin_border
            row_num += 1

    colors = ["FFDDC1", "D3E3FC", "FAF4B7", "D4E2D4"]
    for row in ws.iter_rows(min_row=3, max_row=row_num - 1, min_col=3, max_col=6):
        for idx, cell in enumerate(row):
            cell.fill = PatternFill(start_color=colors[idx], end_color=colors[idx], fill_type="solid")
            cell.border = thin_border

    row_num += 2
    sw_class_headers = ["1. Sınıf", "2. Sınıf", "3. Sınıf"]
    ws.merge_cells(f"A{row_num}:B{row_num}")
    ws[f"A{row_num}"] = "Bölüm"
    ws[f"A{row_num}"].font = Font(bold=True)
    ws[f"A{row_num}"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"A{row_num + 1}:B{row_num + 1}")
    ws[f"A{row_num + 1}"] = "Gün/Saatler"
    ws[f"A{row_num + 1}"].font = Font(bold=True)
    ws[f"A{row_num + 1}"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"C{row_num}:E{row_num}")
    ws[f"C{row_num}"] = "Yazılım Mühendisliği"
    ws[f"C{row_num}"].font = Font(bold=True)
    ws[f"C{row_num}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"C{row_num}"].border = thin_border

    top_border = Border(top=Side(style="thin"), right=Side(style="thin"))
    for col in range(1, 6):
        ws.cell(row=60, column=col).border = top_border

    # Sınıf başlıklarını ekleme
    for idx, header in enumerate(sw_class_headers, start=3):
        cell = ws.cell(row=row_num + 1, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Gün ve saatleri ekleme
    row_num += 2
    for day in days:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + len(time_slots) - 1, end_column=1)
        cell = ws.cell(row=row_num, column=1, value=day)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = thin_border

        for time_slot in time_slots:
            cell = ws.cell(row=row_num, column=2, value=time_slot)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)
            cell.border = thin_border
            row_num += 1

    # Yazılım Mühendisliği için hücreleri renklendirme
    for row in ws.iter_rows(min_row=row_num - len(days) * len(time_slots), max_row=row_num - 1, min_col=3, max_col=5):
        for idx, cell in enumerate(row):
            cell.fill = PatternFill(start_color=colors[idx], end_color=colors[idx], fill_type="solid")
            cell.border = thin_border

    # Excel dosyasını kaydet
    # wb.save("Ders_Programi.xlsx")
    time_slots = ["09:00-10:00", "10:00-11:00", "11:00-12:00",
                  "12:00-13:00", "13:00-14:00", "14:00-15:00", "15:00-16:00",
                  "16:00-17:00", "17:00-18:00", "18:00-19:00", "19:00-20:00", "20:00-21:00"]

    # (Burada senin oluşturduğun kodu aynı şekilde yerleştirebilirsin...)

    dosya_adi = "Ders_Programi.xlsx"
    dosya_yolu = os.path.join(settings.MEDIA_ROOT, dosya_adi)
    generate_schedule_excel(dosya_yolu)

    # Excel içeriğini oku ve listeye aktar
    workbook = openpyxl.load_workbook(dosya_yolu)
    sheet = workbook.active

    excel_tablo = []
    for row in sheet.iter_rows(values_only=True):
        excel_tablo.append(row)

    excel_tablo = normalize_table(excel_tablo)

    # URL ve tabloyu template'e gönder
    context = {
        "excel_dosyasi_var": True,
        "excel_url": settings.MEDIA_URL + dosya_adi,
        "excel_tablo": excel_tablo,
    }

    return render(request, "CourseSchedule/programolustur.html", context)


def tabloyu_getir():
    path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
    wb = load_workbook(path)
    sheet = wb.active
    tablo = []

    onceki_gun = ""

    for row in sheet.iter_rows(values_only=True):
        temiz_satir = []
        for i, cell in enumerate(row):
            if cell is None:
                temiz_satir.append("")
            else:
                temiz_satir.append(cell)

        # Eğer ilk hücre (gün) boşsa, önceki günü yaz
        if temiz_satir[0] == "":
            temiz_satir[0] = onceki_gun
        else:
            onceki_gun = temiz_satir[0]

        tablo.append(temiz_satir)

    return tablo


def saat_araligini_parcala(saat_str):
    saatler = []
    try:
        baslangic, bitis = saat_str.split(",")
        basla = datetime.strptime(baslangic.strip(), "%H:%M")
        bitis = datetime.strptime(bitis.strip(), "%H:%M")
        while basla <= bitis:
            saatler.append(basla.strftime("%H:%M"))
            basla += timedelta(hours=1)
    except Exception as e:
        print("Format Hatası:", saat_str, e)
    return saatler


def dersi_excelden_sil(ders_adi):
    path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
    wb = load_workbook(path)
    ws = wb.active

    silinen = False
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                # Hücreyi satırlara ayır
                satirlar = cell.value.strip().split("\n")
                ilk_satir = satirlar[0].strip() if satirlar else ""

                print(f"Karşılaştır: '{ders_adi}' vs '{ilk_satir}'")
                if ilk_satir == ders_adi.strip():
                    cell.value = None
                    silinen = True

    if silinen:
        wb.save(path)
        print(f"{ders_adi} başarıyla silindi.")
    else:
        print(f"{ders_adi} Excel'de bulunamadı.")


def hucre_hoca_dersi_var_mi(cell_value, ogretim_adi):
    if not cell_value or not isinstance(cell_value, str):
        return False
    satirlar = cell_value.strip().split("\n")
    for satir in satirlar:
        if ogretim_adi.lower() in satir.lower():
            return True
    return False


@csrf_exempt
def common_courses(request):
    query = """
    SELECT DISTINCT d1.ders_adi, d1.ders_kodu
    FROM CourseSchedule_dersler d1
    JOIN CourseSchedule_dersler d2
    ON d1.ders_adi = d2.ders_adi
    AND d1.ders_kodu = d2.ders_kodu
    WHERE (d1.bolum_id = 1 AND d2.bolum_id = 2)
    OR (d1.bolum_id = 2 AND d2.bolum_id = 1)"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        results = cursor.fetchall()
    return results


def excel_ders_atamasi_yap(request, ders_id, ders_adi, ogretim_adi, secimler, bolum, sinif):
    path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
    wb = load_workbook(path)
    ws = wb.active

    saat_satir_map = {
        "09:00": 3, "10:00": 4, "11:00": 5, "12:00": 6,
        "13:00": 7, "14:00": 8, "15:00": 9,
        "16:00": 10, "17:00": 11,
        "18:00": 12, "19:00": 13, "20:00": 14
    }

    gun_satir_offset_map = {
        "Pazartesi": 0,
        "Salı": 12,
        "Çarşamba": 24,
        "Perşembe": 36,
        "Cuma": 48
    }

    yazilim_offset = {
        "Pazartesi": 67,
        "Salı": 79,
        "Çarşamba": 91,
        "Perşembe": 103,
        "Cuma": 115
    }

    bm_sutun_map = {1: 3, 2: 4, 3: 5, 4: 6}
    ysm_sutun_map = {1: 3, 2: 4, 3: 5}

    basari_sayisi = 0

    for secim in secimler:
        try:
            gun, saat = secim.split("|")
            saat = saat.strip()
            satir_index = saat_satir_map.get(saat)

            if bolum == "Bilgisayar Mühendisliği":
                satir = satir_index + gun_satir_offset_map.get(gun, 0)
                sutun = bm_sutun_map.get(sinif)
            elif bolum == "Yazılım Mühendisliği":
                satir = yazilim_offset.get(gun, 0) + (satir_index - 3)
                sutun = ysm_sutun_map.get(sinif)
            else:
                continue

            # Formdan gelen seçilmiş dersliği oku
            derslik_key = f"derslik_{gun}_{saat}"
            secilen_derslik = request.POST.get(derslik_key)

            if secilen_derslik:
                yazilacak_veri = f"{ders_adi}\n{ogretim_adi} ({secilen_derslik})"
                ws.cell(row=satir, column=sutun).value = yazilacak_veri
                basari_sayisi += 1
            else:
                print(f"[UYARI] Formdan seçili derslik alınamadı: {derslik_key}")

        except Exception as e:
            print(f"[HATA] Atama sırasında hata oluştu: {secim} - {e}")

    wb.save(path)
    return f"{ders_adi} dersi {basari_sayisi} saate başarıyla atandı."


def uygun_derslikleri_getir(satir_index, ders_id):
    path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
    wb = load_workbook(path)
    ws = wb.active

    # Tüm tanımlı derslikleri ve kapasitelerini al
    tum_derslikler = Derslikler.objects.all()
    kullanilan_derslikler = set()

    # O satırdaki tüm hücrelerde geçen derslikleri topla
    for col in range(3, ws.max_column + 1):
        hucre = ws.cell(row=satir_index, column=col)
        if isinstance(hucre.value, str):
            for derslik in tum_derslikler:
                if f"({derslik.derslik_id})" in hucre.value:
                    kullanilan_derslikler.add(derslik.derslik_id)

    # O dersi alan öğrenci sayısını hesapla
    ogrenci_sayisi = OgrenciDers.objects.filter(ders_id=ders_id).count()

    # Uygun derslikleri filtrele
    uygunlar = []
    for derslik in tum_derslikler:
        if derslik.derslik_id not in kullanilan_derslikler and derslik.kapasite >= ogrenci_sayisi:
            uygunlar.append(derslik.derslik_id)

    # Log yazdır
    print("Öğrenci sayısı:", ogrenci_sayisi)
    print("Kullanılan derslikler:", kullanilan_derslikler)
    print("Uygunlar:", uygunlar)

    return uygunlar


def manuel_atama(request):
    #dersler = Dersler.objects.all()
    tum_dersler = Dersler.objects.all()
    görüntülenecek_dersler = {}
    for ders in tum_dersler:
        key = (ders.ders_adi.strip(), ders.ders_kodu.strip())
        if key not in görüntülenecek_dersler:
            görüntülenecek_dersler[key] = ders
    dersler = list(görüntülenecek_dersler.values())
    secilen_ders = None
    mevcut_saat = ""
    mevcut_derslik = ""
    alternatifler_bm = []  # Bilgisayar Mühendisliği için alternatifler
    alternatifler_ysm = []  # Yazılım Mühendisliği için alternatifler

    def get_surname(full_name):
        return full_name.strip().split()[-1].lower()

    if request.method == "POST":
        form = ManuelAtamaForm()
        ders_id = request.POST.get("ders_id")
        if ders_id:
            ders = get_object_or_404(Dersler, pk=ders_id)
            ders_adi = ders.ders_adi.strip()
            ogretim_adi = ders.ogretim_uyesi.ogretim_gorevlisi.strip()
            bolum = ders.bolum.bolum_adi
            sinif = ders.sinif

            # Silme işlemi
            if "sil" in request.POST:
                dersi_excelden_sil(ders_adi)
                messages.success(request, f"{ders_adi} Excel’den silindi.")
                return redirect("manuel_atama")

            # Atama işlemi (BM veya YSM)
            if "atama_bm" in request.POST or "atama_ysm" in request.POST:
                if "atama_bm" in request.POST:
                    secimler = request.POST.getlist("secimler_bm")
                    bolum = "Bilgisayar Mühendisliği"
                else:
                    secimler = request.POST.getlist("secimler_ysm")
                    bolum = "Yazılım Mühendisliği"

                mesaj = excel_ders_atamasi_yap(request, ders_id, ders_adi, ogretim_adi, secimler, bolum, sinif)
                messages.success(request, mesaj)
                return redirect("manuel_atama")


    else:
        form = ManuelAtamaForm()
        ders_id = request.GET.get("ders_id")
        if ders_id:
            secilen_ders = get_object_or_404(Dersler, pk=ders_id)
            mevcut_saat = secilen_ders.zorunlu_saat or ""
            mevcut_derslik = secilen_ders.get_statu_display()
            ogr = secilen_ders.ogretim_uyesi
            teacher = ogr.ogretim_gorevlisi.strip()
            surname = get_surname(teacher)
            sinif = secilen_ders.sinif

            # Excel dosyasını aç
            path = os.path.join(settings.MEDIA_ROOT, 'Ders_Programi.xlsx')
            wb = load_workbook(path)
            ws = wb.active

            gun_saat_map = {
                "Pazartesi": ogr.pazartesi,
                "Salı": ogr.sali,
                "Çarşamba": ogr.carsamba,
                "Perşembe": ogr.persembe,
                "Cuma": ogr.cuma
            }

            saat_satir_map = {
                "09:00": 3, "10:00": 4, "11:00": 5, "12:00": 6,
                "13:00": 7, "14:00": 8, "15:00": 9,
                "16:00": 10, "17:00": 11,
                "18:00": 12, "19:00": 13, "20:00": 14
            }

            gun_satir_offset_map = {
                "Pazartesi": 0,
                "Salı": 12,
                "Çarşamba": 24,
                "Perşembe": 36,
                "Cuma": 48
            }

            sinif_sutun_map = {1: 3, 2: 4, 3: 5, 4: 6}
            kendi_sinif_sutun = sinif_sutun_map.get(sinif)

            # Ortak dersler sorgulaması
            ortak_dersler = common_courses(request)

            # Ortak dersler arasında secilen_ders bulunuyorsa, saatleri al

            if any(ders_adi == secilen_ders.ders_adi for ders_adi, _ in ortak_dersler):
                # 1. Bilgisayar Mühendisliği alternatif saatleri
                for gun, saat_str in gun_saat_map.items():
                    if not saat_str:
                        continue
                    for aralik in saat_str.split(";"):
                        for saat in saat_araligini_parcala(aralik.strip()):
                            temel_satir = saat_satir_map.get(saat)
                            if not (temel_satir and kendi_sinif_sutun):
                                continue

                            satir = temel_satir + gun_satir_offset_map.get(gun, 0)

                            # 1. Adım: Kendi sınıfında boş mu?
                            hucre = ws.cell(row=satir, column=kendi_sinif_sutun)
                            if hucre.value and str(hucre.value).strip():
                                print(f"[DOLU] Kendi sınıfında hücre dolu: {gun} {saat}")
                                continue

                            # 2. Adım: Aynı anda başka sınıfta dersi var mı?
                            diger_sinif_sutunlari = [col for col in sinif_sutun_map.values() if
                                                     col != kendi_sinif_sutun]
                            cakisma_var = False
                            for col in diger_sinif_sutunlari:
                                hucre_diger = ws.cell(row=satir, column=col)
                                if hucre_hoca_dersi_var_mi(hucre_diger.value, teacher):
                                    print(f"[ÇAKIŞMA] {teacher} aynı anda başka sınıfta ders veriyor: {gun} {saat}")
                                    cakisma_var = True
                                    break
                            if cakisma_var:
                                continue

                            uygun_derslikler = uygun_derslikleri_getir(satir, secilen_ders.pk)

                            # Uygunsa listeye ekle
                            alternatifler_bm.append({
                                "gun": gun,
                                "saat": saat,
                                "derslik": uygun_derslikler
                            })

                # 2. Yazılım Mühendisliği alternatif saatleri
                ym_offset = {"Pazartesi": 67, "Salı": 79, "Çarşamba": 91, "Perşembe": 103, "Cuma": 115}
                ym_sutun_map = {1: 3, 2: 4, 3: 5}
                kendi_sutun = ym_sutun_map.get(sinif)

                for gun, saat_str in gun_saat_map.items():
                    if not saat_str:
                        continue

                    for aralik in saat_str.split(";"):
                        for saat in saat_araligini_parcala(aralik.strip()):
                            saat_indeksi = saat_satir_map.get(saat)
                            if saat_indeksi is None:
                                continue

                            # YZM çizelgesi satır indeksini hesapla
                            satir = ym_offset[gun] + (saat_indeksi - 3)

                            # 1. Adım: YZM çizelgesinde o saatte ders var mı?
                            hucre_yzm = ws.cell(row=satir, column=kendi_sutun)
                            if hucre_yzm.value and str(hucre_yzm.value).strip():
                                print(f"[YAZILIM DOLU] Yazılım Müh. çizelgesinde hücre dolu: {gun} {saat}")
                                continue

                            # 2. Adım: Aynı anda diğer sınıflarda bu öğretim üyesinin dersi var mı?
                            cakisma_var = False
                            for diger_sinif, diger_sutun in sinif_sutun_map.items():
                                if diger_sutun == kendi_sinif_sutun:
                                    continue
                                hucre_diger = ws.cell(row=satir, column=diger_sutun)
                                if hucre_hoca_dersi_var_mi(hucre_diger.value, teacher):
                                    print(f"[ÇAKIŞMA] {teacher} aynı anda başka sınıfta ders veriyor: {gun} {saat}")
                                    cakisma_var = True
                                    break

                            if cakisma_var:
                                continue

                            uygun_derslikler = uygun_derslikleri_getir(satir, secilen_ders.pk)

                            # Uygunsa listeye ekle
                            alternatifler_ysm.append({
                                "gun": gun,
                                "saat": saat,
                                "derslik": uygun_derslikler
                            })


            # Alternatif saatler yalnızca Bilgisayar Mühendisliği veya Yazılım Mühendisliği için de alınabilir.
            else:
                # Bilgisayar Mühendisliği için alternatif saatlerin belirlenmesi
                if secilen_ders.bolum.bolum_adi == "Bilgisayar Mühendisliği":
                    for gun, saat_str in gun_saat_map.items():
                        if not saat_str:
                            continue
                        for aralik in saat_str.split(";"):
                            for saat in saat_araligini_parcala(aralik.strip()):
                                temel_satir = saat_satir_map.get(saat)
                                if not (temel_satir and kendi_sinif_sutun):
                                    continue

                                satir = temel_satir + gun_satir_offset_map.get(gun, 0)

                                # 1. Adım: Kendi sınıfında boş mu?
                                hucre = ws.cell(row=satir, column=kendi_sinif_sutun)
                                if hucre.value and str(hucre.value).strip():
                                    print(f"[DOLU] Kendi sınıfında hücre dolu: {gun} {saat}")
                                    continue

                                # 2. Adım: Aynı anda başka sınıfta dersi var mı?
                                diger_sinif_sutunlari = [col for col in sinif_sutun_map.values() if
                                                         col != kendi_sinif_sutun]
                                cakisma_var = False
                                for col in diger_sinif_sutunlari:
                                    hucre_diger = ws.cell(row=satir, column=col)
                                    if hucre_hoca_dersi_var_mi(hucre_diger.value, teacher):
                                        print(f"[ÇAKIŞMA] {teacher} aynı anda başka sınıfta ders veriyor: {gun} {saat}")
                                        cakisma_var = True
                                        break
                                if cakisma_var:
                                    continue

                                uygun_derslikler = uygun_derslikleri_getir(satir, secilen_ders.pk)

                                # Uygunsa listeye ekle
                                alternatifler_bm.append({
                                    "gun": gun,
                                    "saat": saat,
                                    "derslik": uygun_derslikler
                                })


                # Yazılım Mühendisliği için alternatif saatlerin belirlenmesi
                elif secilen_ders.bolum.bolum_adi == "Yazılım Mühendisliği":
                    ym_offset = {"Pazartesi": 67, "Salı": 79, "Çarşamba": 91, "Perşembe": 103, "Cuma": 115}
                    ym_sutun_map = {1: 3, 2: 4, 3: 5}
                    kendi_sutun = ym_sutun_map.get(sinif)

                    if not kendi_sutun:
                        print("Sınıf bilgisi eksik")
                    else:
                        for gun, saat_str in gun_saat_map.items():
                            if not saat_str:
                                continue
                            for aralik in saat_str.split(";"):
                                for saat in saat_araligini_parcala(aralik.strip()):
                                    saat_indeksi = saat_satir_map.get(saat)
                                    if saat_indeksi is None:
                                        continue

                                    satir = ym_offset[gun] + (saat_indeksi - 3)

                                    # 1. Adım: Yazılım Mühendisliği çizelgesinde o saatte ders var mı?
                                    hucre_yzm = ws.cell(row=satir, column=kendi_sutun)
                                    if hucre_yzm.value and str(hucre_yzm.value).strip():
                                        print(f"[YAZILIM DOLU] Yazılım Müh. çizelgesinde hücre dolu: {gun} {saat}")
                                        continue

                                    # 2. Adım: Aynı anda diğer sınıflarda bu öğretim üyesinin dersi var mı?
                                    cakisma_var = False
                                    for diger_sinif, diger_sutun in ym_sutun_map.items():
                                        if diger_sutun == kendi_sutun:
                                            continue
                                        hucre_diger = ws.cell(row=satir, column=diger_sutun)
                                        if hucre_hoca_dersi_var_mi(hucre_diger.value, teacher):
                                            print(
                                                f"[ÇAKIŞMA] {teacher} aynı anda başka sınıfta ders veriyor: {gun} {saat}")
                                            cakisma_var = True
                                            break
                                    if cakisma_var:
                                        continue

                                    uygun_derslikler = uygun_derslikleri_getir(satir, secilen_ders.pk)

                                    # Uygunsa listeye ekle
                                    alternatifler_ysm.append({
                                        "gun": gun,
                                        "saat": saat,
                                        "derslik": uygun_derslikler
                                    })

    return render(request, "CourseSchedule/manuel_atama.html", {
        "form": form,
        "dersler": dersler,
        "secilen_ders": secilen_ders,
        "mevcut_saat": mevcut_saat,
        "mevcut_derslik": mevcut_derslik,
        "alternatifler_bm": alternatifler_bm,
        "alternatifler_ysm": alternatifler_ysm,
        "excel_tablo": tabloyu_getir(),
    })











